import os
from os.path import basename
import re
from flask_login.utils import logout_user
import pandas as pd
import openpyxl
import base64
import pdfkit
from datetime import date
from zipfile import ZipFile
from openpyxl.utils.dataframe import dataframe_to_rows
from flask import Flask,render_template, request, flash, redirect, send_file, send_from_directory, safe_join, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, current_user, login_user
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from config import UPLOAD_FOLDER, ALLOWED_EXTENSIONS, CHOICE_CODES, SECRET_KEY, ENSAM_CODES
from models import Results, LPCasa, LPMeknes, LPRabat, LACasa, LAMeknes, LARabat, User, LoginForm, db
 
app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root:@localhost/assignementplatform' 
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SECRET_KEY'] = SECRET_KEY
login = LoginManager(app)
db.init_app(app)

def allowedFile(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def fileExtension(filename):
    return filename.rsplit('.', 1)[1].lower()

def toPdf(name, typ, listesPrincipales, listesAttentes):
    with open('assets/logo'+name.capitalize()+'.png', "rb") as image_file:
        b64Image = base64.b64encode(image_file.read()).decode()
    
    # header = open('assets/'+name+'Header.html', 'r')
    # if not os.path.isdir('html'):
    #     os.mkdir('html')
    # with open('html/'+name+'.html', 'w') as file:
    #     for line in header:
    #         file.write(line.replace('date', str(date.today().year-1)+'-'+str(date.today().year)).replace('$b64Img$', b64Image))
    #     file.write(listesPrincipales[name][['cne','nom']].to_html(index=False, table_id=name+'Table').replace('border="1"', ''))
    #     file.write('</body>')
    # header.close()
    
    if typ == 'LP':
        header = open('assets/'+name+'Header.html', 'r')
        html=''
        for line in header:
            html += line.replace('date', str(date.today().year-1)+'-'+str(date.today().year)).replace('$b64Img$', b64Image)
        html += listesPrincipales[name][['cne','nomPrenom']].to_html(index=False, table_id=name+'Table').replace('border="1"', '')
        html+='</body>'
        header.close()
    elif typ =='LA':
        header = open('assets/'+name+'AttenteHeader.html', 'r')
        html=''
        for line in header:
            html += line.replace('date', str(date.today().year-1)+'-'+str(date.today().year)).replace('$b64Img$', b64Image)
        html += listesAttentes[name][['cne','nomPrenom']].to_html(index=False, table_id=name+'Table').replace('border="1"', '')
        html+='</body>'
        header.close()
    options = {
        "enable-local-file-access": None,
        'quiet': ''
    }
    
    if not os.path.isdir('output'):
        os.mkdir('output')
    pdfkit.from_string(html, 'output/'+typ+'_'+name+'.pdf', options=options, css='assets/style.css')

@login.user_loader
def load_user(id):
    return User.query.get(id)

@app.route('/')
def index():
    if current_user.is_authenticated:
        try:
            results = Results.query.order_by(Results.moyenne.desc()).all()
            lp_casa = LPCasa.query.order_by(LPCasa.nomPrenom).all()
            lp_meknes = LPMeknes.query.order_by(LPMeknes.nomPrenom).all()
            lp_rabat = LPRabat.query.order_by(LPRabat.nomPrenom).all()
            la_casa = LACasa.query.order_by(LACasa.moyenne.desc()).all()
            la_meknes = LAMeknes.query.order_by(LAMeknes.moyenne.desc()).all()
            la_rabat = LARabat.query.order_by(LARabat.moyenne.desc()).all()
        except Exception as e:
            error_text = "<p>The error:<br>" + str(e) + "</p>"
            hed = '<h1>Something is broken.</h1>'
            return hed + error_text
        if not results:
            results = 'NULL'
        elif not(lp_casa and lp_meknes and lp_rabat):
            lp_casa = 'NULL'
            lp_meknes = 'NULL'
            lp_rabat = 'NULL'
        return render_template('index.html', results=results, lp_casa=lp_casa, lp_meknes=lp_meknes, lp_rabat=lp_rabat, la_casa=la_casa, la_meknes=la_meknes, la_rabat=la_rabat, current_user = current_user)
    else:
        return redirect('/login')

@app.route('/uploadResults', methods=['POST'])
def uploadResults():
    if current_user.is_authenticated:
        if request.method == 'POST':
            if 'results' not in request.files:
                return redirect('/')
            file = request.files['results']
            if file.filename == '':
                return redirect('/')
            if file and allowedFile(file.filename):
                filename = secure_filename(file.filename)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], 'results.'+fileExtension(filename)))
                if fileExtension(filename) == 'xlsx':
                    resultsDf = pd.read_excel(os.path.join(app.config['UPLOAD_FOLDER'], 'results.xlsx'))
                elif fileExtension(filename) == 'csv':
                    resultsDf = pd.read_csv(os.path.join(app.config['UPLOAD_FOLDER'], 'results.csv'), names=['CNE', 'NOM', 'Choix 1', 'Choix 2', 'Choix 3', 'Moyenne'])
                resultsDf = resultsDf[['CNE', 'NOM', 'Choix 1', 'Choix 2', 'Choix 3', 'FILIERE', 'Note Maths', 'Note Physique', 'Moyenne']]
                resultsDf = resultsDf.rename(columns={'CNE':'cne', 'NOM':'nomPrenom', 'Choix 1':'choix1', 'Choix 2':'choix2', 'Choix 3':'choix3', 'FILIERE': 'filiere', 'Note Maths':'noteMaths', 'Note Physique':'notePhysique', 'Moyenne':'moyenne'})
                resultsDf.to_sql('results', con=db.engine, index=False, if_exists='replace')
                return redirect('/')

@app.route('/confirmStudents', methods=['POST'])
def confirmStudents():
    if current_user.is_authenticated:
        if request.form['submit'] == 'casa':
            lp = pd.read_sql('SELECT * FROM lp_casa', con=db.engine)
            confirmed = request.form.getlist('confirmedCasa')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('lp_casa', con=db.engine, index=False, if_exists='replace')
        elif request.form['submit'] == 'meknes':
            lp = pd.read_sql('SELECT * FROM lp_meknes', con=db.engine)
            confirmed = request.form.getlist('confirmedMeknes')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('lp_meknes', con=db.engine, index=False, if_exists='replace')
        elif request.form['submit'] == 'rabat':
            lp = pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)
            confirmed = request.form.getlist('confirmedRabat')
            lp['confirmed'] = False
            lp.loc[lp.cne.isin(confirmed), 'confirmed'] = True
            lp.to_sql('lp_rabat', con=db.engine, index=False, if_exists='replace')
    
    return redirect('/genererLA')

@app.route('/genererLA', methods=['GET', 'POST'])
def genererLA():
    if current_user.is_authenticated:
        listesAttentes = {'casa':[], 'meknes':[], 'rabat':[]}
        listesPrincipales = {'casa':pd.read_sql('SELECT * FROM lp_casa', con=db.engine), 'meknes':pd.read_sql('SELECT * FROM lp_meknes', con=db.engine), 'rabat':pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)}
        listesReaffectation = {}
        results = pd.read_sql('SELECT * FROM results', con=db.engine)
        choiceCodes = CHOICE_CODES
        AVAILABLE_PLACES = {'casa':0, 'meknes':0, 'rabat':0}
        
        maxAdmis = 0
        origin=[]
        for key in listesPrincipales:
            listesReaffectation[key] = listesPrincipales[key][listesPrincipales[key]['confirmed']==True]
            listesReaffectation[key]['origin'] = key
            maxAdmis += len(listesPrincipales[key].index)
        
        for key in AVAILABLE_PLACES:
            AVAILABLE_PLACES[key] = len(listesPrincipales[key].index)-len(listesReaffectation[key].index) 
        
        #do the reallocation
            #remove all confirmed members that were accepted in the LP
        for key in listesReaffectation:
            for index, row in listesReaffectation[key].iterrows():
                if row['choix1'] == ENSAM_CODES[key]:
                    listesReaffectation[key] = listesReaffectation[key].drop(index)
        reaffectedDF = pd.concat([listesReaffectation['casa'], listesReaffectation['meknes'], listesReaffectation['rabat']], ignore_index=True)
        reaffectedDF = reaffectedDF.sort_values(by=['moyenne'], ascending=False)
        
        print(AVAILABLE_PLACES)
        for index, row in reaffectedDF.iterrows():
                if AVAILABLE_PLACES[choiceCodes[row['choix1']]] > 0:
                    listesAttentes[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
                    AVAILABLE_PLACES[choiceCodes[row['choix1']]]-=1
                    AVAILABLE_PLACES[row['origin']]+=1            
                elif AVAILABLE_PLACES[choiceCodes[row['choix2']]]>0:            
                    listesAttentes[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
                    AVAILABLE_PLACES[choiceCodes[row['choix2']]]-=1   
                    AVAILABLE_PLACES[row['origin']]+=1            
                elif AVAILABLE_PLACES[choiceCodes[row['choix3']]]>0:             
                    listesAttentes[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
                    AVAILABLE_PLACES[choiceCodes[row['choix3']]]-=1 
                    AVAILABLE_PLACES[row['origin']]+=1            
                else:
                    break
        
        print(AVAILABLE_PLACES)
        print(reaffectedDF)
        
        listeAttenteDf = results.drop(index=results.index[:maxAdmis],axis=0)
        for index, row in listeAttenteDf.iterrows():
            listesAttentes[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
        
        for key in listesAttentes:
                listesAttentes[key]=pd.DataFrame(listesAttentes[key])
                listesAttentes[key].to_sql('la_'+key, con=db.engine, index=False, if_exists='replace')
        
        return redirect('/')

@app.route('/genererLP', methods=['POST'])
def genererLP():
    if current_user.is_authenticated:
        if request.method == 'POST':
            colNames = ['cne', 'nomPrenom', 'choix1', 'choix2', 'choix3', 'filiere', 'noteMaths', 'notePhysique', 'moyenne']
            AVAILABLE_PLACES = {'casa':int(request.form.get('CASA_MAX_PLACES')), 'meknes':int(request.form.get('MEKNES_MAX_PLACES')), 'rabat':int(request.form.get('RABAT_MAX_PLACES'))}
            listesPrincipales = {'casa':[], 'meknes':[], 'rabat':[]}
            indexes={'casa':0, 'meknes':0, 'rabat':0} 
            choiceCodes=CHOICE_CODES
            
            results = pd.read_sql('SELECT * FROM results', con=db.engine)
            
            results.sort_values(by=['moyenne'], inplace=True, ascending=False)
            results.dropna(subset=['cne', 'nomPrenom', 'moyenne'], inplace=True)
            nbEtudiants = len(results.index)
            
            for index, row in results.iterrows():
                if AVAILABLE_PLACES[choiceCodes[row['choix1']]] > 0:
                    listesPrincipales[choiceCodes[row['choix1']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
                    AVAILABLE_PLACES[choiceCodes[row['choix1']]]-=1
                    indexes[choiceCodes[row['choix1']]]+=1                
                elif AVAILABLE_PLACES[choiceCodes[row['choix2']]]>0:            
                    listesPrincipales[choiceCodes[row['choix2']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
                    AVAILABLE_PLACES[choiceCodes[row['choix2']]]-=1
                    indexes[choiceCodes[row['choix2']]]+=1    
                elif AVAILABLE_PLACES[choiceCodes[row['choix3']]]>0:             
                    listesPrincipales[choiceCodes[row['choix3']]].append({'cne': row['cne'], 'nomPrenom': row['nomPrenom'], 'choix1': row['choix1'], 'choix2': row['choix2'], 'choix3': row['choix3'], 'filiere':row['filiere'], 'noteMaths':row['noteMaths'], 'notePhysique':row['notePhysique'], 'moyenne':row['moyenne']})
                    AVAILABLE_PLACES[choiceCodes[row['choix3']]]-=1
                    indexes[choiceCodes[row['choix3']]]+=1  
                else:
                    break
            
            for key in listesPrincipales:
                listesPrincipales[key]=pd.DataFrame(listesPrincipales[key])
                listesPrincipales[key]['confirmed'] = False
                listesPrincipales[key].to_sql('lp_'+key, con=db.engine, index=False, if_exists='replace')

            return redirect('/')

@app.route('/downloadFiles')
def downloadFiles():
    if current_user.is_authenticated:
        listesPrincipales = {'casa':pd.read_sql('SELECT * FROM lp_casa', con=db.engine), 'meknes':pd.read_sql('SELECT * FROM lp_meknes', con=db.engine), 'rabat':pd.read_sql('SELECT * FROM lp_rabat', con=db.engine)}
        listesAttentes = {'casa':pd.read_sql('SELECT * FROM la_casa', con=db.engine), 'meknes':pd.read_sql('SELECT * FROM la_meknes', con=db.engine), 'rabat':pd.read_sql('SELECT * FROM la_rabat', con=db.engine)}
        results = pd.read_sql('SELECT * FROM results', con=db.engine)
        
        nbEtudiants = len(results.index)
        
        wb = openpyxl.load_workbook('output/Resultats.xlsx')
        ws = wb['resultats']
        ws.delete_rows(0,4*nbEtudiants)
        for row in dataframe_to_rows(results.loc[:,~results.columns.str.match("Unnamed")], index=False):
            ws.append(row)

        maxCol = ws.max_column
        ws.cell(1, maxCol+1).value = 'Affectation Liste Principale'
        ws.cell(1, maxCol+2).value = 'Affectation Liste d\'Attente'
        
        
        for i in range(2,nbEtudiants+2):
            flag=1
            for key in listesPrincipales:
                if ws['A'+str(i)].value in listesPrincipales[key].cne.values:
                    ws.cell(i, maxCol+1).value = key.capitalize()
                    flag=0
            if flag:
                ws.cell(i, maxCol+1).value = ''

        for i in range(2,nbEtudiants+2):
            flag=1
            for key in listesAttentes:
                if ws['A'+str(i)].value in listesAttentes[key].cne.values:
                    ws.cell(i, maxCol+2).value = key.capitalize()
                    flag = 0
            if flag:
                ws.cell(i, maxCol+2).value = ''

        for key in listesPrincipales:               
            name = 'LP_'+key
            if name not in wb.sheetnames:
                wb.create_sheet(name)
                ws = wb[name]
            else:
                ws = wb[name]
                ws.delete_rows(2,2*nbEtudiants)
            
            for row in dataframe_to_rows(listesPrincipales[key].drop('confirmed', axis=1), index=False):
                ws.append(row)
                        
            listesPrincipales[key].sort_values(by=['nomPrenom'], inplace=True)
            toPdf(key, 'LP', listesPrincipales, listesAttentes) 
        
        
        for key in listesAttentes:               
            name = 'LA_'+key
            if name not in wb.sheetnames:
                wb.create_sheet(name)
                ws = wb[name]
            else:
                ws = wb[name]
                ws.delete_rows(2,2*nbEtudiants)
            
            for row in dataframe_to_rows(listesAttentes[key], index=False):
                ws.append(row)
            
            toPdf(key, 'LA', listesPrincipales, listesAttentes) 
        
        ws = wb['resultats']
        for cell in ws["1:1"]:
            cell.font = openpyxl.styles.Font(color='00000000', bold=True, size='12') 
            cell.alignment = openpyxl.styles.alignment.Alignment(horizontal = 'center', vertical ='center')
        
        for row in ws.iter_rows():
            for cell in row:
                cell.border = openpyxl.styles.borders.Border(left=openpyxl.styles.borders.Side(style='thin'), right=openpyxl.styles.borders.Side(style='thin'), top=openpyxl.styles.borders.Side(style='thin'), bottom=openpyxl.styles.borders.Side(style='thin'))
        
        wb.save('output/Resultats.xlsx')
        return redirect('/zipnsend')

@app.route('/zipnsend')
def zipnsend():
    if current_user.is_authenticated:
        with ZipFile('zip/output.zip', 'w') as zipObj:
            for folderName, subfolders, filenames in os.walk('output'):
                for filename in filenames:
                    #create complete filepath of file in directory
                    filePath = os.path.join(folderName, filename)
                    # Add file to zip
                    zipObj.write(filePath, basename(filePath))
        
        return redirect('/zip/output.zip')

@app.route("/zip/output.zip")
def getFile():
    try:
        return send_from_directory('zip', 'output.zip', as_attachment=True)
    except FileNotFoundError:
        abort(404)       
        
        

@app.route('/genererpasswords')
def passwd():
    passwordCasa = 'a47CyM&z4UkD'
    passwordMeknes = 'HkT6Hy!ZrV4!'
    passwordRabat = '8j5$TzCpD^s@'
    passwordAdmin = 'admin'
    Casa = User(username='ENSAM_Casa', password=generate_password_hash(passwordCasa))
    Meknes = User(username='ENSAM_Meknes', password=generate_password_hash(passwordMeknes))
    Rabat = User(username='ENSAM_Rabat', password=generate_password_hash(passwordRabat))
    Admin = User(username='admin', password=generate_password_hash(passwordAdmin))
    db.session.add(Casa)
    db.session.add(Meknes)
    db.session.add(Rabat)
    db.session.add(Admin)
    db.session.commit()
    return redirect('/')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        logout_user()
        return redirect('/')
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user is None or not user.check_password(form.password.data):
            flash('Invalid username or password')
            return redirect('/login')
        login_user(user)
        return redirect('/')
    return render_template('login.html', title='Sign In', form=form)

if __name__ == '__main__':
    app.run(host='localhost', port=5000, debug=True)